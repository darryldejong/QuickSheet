from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment

# kleurcodes
rood = "\033[91m"
groen = "\033[92m"
blauw = "\033[94m"
paars = "\033[95m"
cyaan = "\033[94m"
geel = "\033[93m"
reset = "\033[0m"

def clear_screen():
    print("\033c", end="")

def set_column_width(ws, column, width):
    ws.column_dimensions[column].width = width

def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height

def sort_data(data, sort_order='ascending'):
    header = data[0]
    body = data[1:]

    if sort_order == 'ascending':
        sorted_body = sorted(body, key=lambda x: x[0])
    else:
        sorted_body = sorted(body, key=lambda x: x[0], reverse=True)

    return [header] + sorted_body

def create_excel_table(data, file_name='tabel.xlsx', border_style='thin', col_width='medium', row_height='medium'):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Data'
        
        for row in data:
            ws.append(row)

        if border_style == 'thin':
            border = Border(left=Side(border_style="thin", color="000000"),
                            right=Side(border_style="thin", color="000000"),
                            top=Side(border_style="thin", color="000000"),
                            bottom=Side(border_style="thin", color="000000"))
        elif border_style == 'thick':
            border = Border(left=Side(border_style="thick", color="000000"),
                            right=Side(border_style="thick", color="000000"),
                            top=Side(border_style="thick", color="000000"),
                            bottom=Side(border_style="thick", color="000000"))
        elif border_style == 'dotted':
            border = Border(left=Side(border_style="dotted", color="000000"),
                            right=Side(border_style="dotted", color="000000"),
                            top=Side(border_style="dotted", color="000000"),
                            bottom=Side(border_style="dotted", color="000000"))
        else:
            print(f"{rood}❌ Ongeldige randstijl gekozen. Er wordt een dunne rand gebruikt als standaard.{reset}")
            border = Border(left=Side(border_style="thin", color="000000"),
                            right=Side(border_style="thin", color="000000"),
                            top=Side(border_style="thin", color="000000"),
                            bottom=Side(border_style="thin", color="000000"))

        for row in ws.iter_rows(min_row=1, max_col=len(data[0]), max_row=len(data)):
            for cell in row:
                cell.border = border

        header_font = Font(size=14)
        header_alignment = Alignment(horizontal="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = header_alignment

        # kolombreedte en rijhoogte
        width_options = {1: 10, 2: 15, 3: 25}
        height_options = {1: 15, 2: 20, 3: 30}

        chosen_col_width = width_options.get(col_width, 15)
        chosen_row_height = height_options.get(row_height, 20)

        for i in range(1, len(data[0]) + 1):
            set_column_width(ws, ws.cell(row=1, column=i).column_letter, chosen_col_width)

        for i in range(1, len(data) + 1):
            set_row_height(ws, i, chosen_row_height)

        wb.save(file_name)
        
        print(f'{groen}✅ Tabel succesvol aangemaakt en opgeslagen als {file_name}{reset}')
    except PermissionError:
        print(f"{rood}❌ Fout bij het aanmaken van het Excel-bestand: Toegang geweigerd voor {file_name}.{reset}")
    except Exception as e:
        print(f"{rood}❌ Fout bij het aanmaken van het Excel-bestand: {e}{reset}")

def get_user_input(num_rows, num_cols):
    rows = []
    for i in range(num_rows):
        while True:
            row_data = input(f"Voer de gegevens in voor rij {i + 1} (spaties tussen de waarden): ")
            if row_data == '':
                print(f"{rood}❌ Invoer mag niet leeg zijn. Probeer het opnieuw.{reset}")
                continue
            try:
                row_data = row_data.split()
                if len(row_data) != num_cols:
                    print(f"{rood}❌ Fout: Voer precies {num_cols} waarden in.{reset}")
                    continue
                rows.append(row_data)
                break
            except Exception as e:
                print(f"{rood}❌ Fout bij het verwerken van de invoer: {e}{reset}")

    return rows

def get_valid_input(prompt, valid_options):
    while True:
        try:
            choice = int(input(prompt))
            if choice in valid_options:
                return choice
            else:
                print(f"{rood}❌ Ongeldige keuze. Kies alstublieft uit de opties: {valid_options}.{reset}")
        except ValueError:
            print(f"{rood}❌ Ongeldige invoer. Voer een getal in.{reset}")

def main():
    while True:
        clear_screen()
        print(f"{cyaan}Excel Tabel Generator{reset}")
        print("Dit script helpt je om gegevens in te voeren en op te slaan in een Excel-bestand.")
        
        try:
            num_cols = int(input("Aantal kolommen: "))
            num_rows = int(input("Aantal rijen: "))
        except ValueError:
            print(f"{rood}❌ Fout: Voer een geldig getal in voor het aantal kolommen en rijen.{reset}")
            continue

        print("Geef de kolomnamen in (spaties tussen de namen): ")
        headers = input().split()
        
        if len(headers) != num_cols:
            print(f"{rood}❌ Fout: Je hebt {len(headers)} kolommen ingevoerd, maar je hebt gevraagd om {num_cols}.{reset}")
        else:
            data = get_user_input(num_rows, num_cols)
            data = [headers] + data

            print("Kies een sorteermethode:")
            print(f"1: {blauw}Van A naar Z{reset}")
            print(f"2: {paars}Van Z naar A{reset}")
            print(f"3: {groen}Geen sortering (standaard){reset}")
            sort_choice = get_valid_input("Kies een optie (1/2/3): ", [1, 2, 3])
            
            if sort_choice == 1:
                data = sort_data(data, sort_order='ascending')
            elif sort_choice == 2:
                data = sort_data(data, sort_order='descending')

            print("Kies een randstijl:")
            print(f"1: {blauw}Dun{reset}")
            print(f"2: {paars}Dik{reset}")
            print(f"3: {groen}Gestippeld{reset}")
            border_style_choice = get_valid_input("Kies een optie (1/2/3): ", [1, 2, 3])

            if border_style_choice == 1:
                border_style = 'thin'
            elif border_style_choice == 2:
                border_style = 'thick'
            elif border_style_choice == 3:
                border_style = 'dotted'

            print("Kies de kolombreedte:")
            print(f"1: {blauw}Klein{reset}")
            print(f"2: {paars}Gemiddeld{reset}")
            print(f"3: {groen}Groot{reset}")
            col_width_choice = get_valid_input("Kies een optie (1/2/3): ", [1, 2, 3])

            print("Kies de rijhoogte:")
            print(f"1: {blauw}Klein{reset}")
            print(f"2: {paars}Gemiddeld{reset}")
            print(f"3: {groen}Groot{reset}")
            row_height_choice = get_valid_input("Kies een optie (1/2/3): ", [1, 2, 3])

            create_excel_table(data, border_style=border_style, col_width=col_width_choice, row_height=row_height_choice)

        choice = input(f"{geel}Druk op 'p' om opnieuw te starten of 'q' om af te sluiten: {reset}").lower()
        if choice == 'q':
            break

if __name__ == '__main__':
    main()
